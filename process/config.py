from typing import Dict, List
from openpyxl import load_workbook

# Global mappings - each contains workbook name -> list of row dictionaries
excel_mappings: Dict[str, List[Dict[str, str]]] = {}


def get_excel_mapping() -> Dict[str, List[Dict[str, str]]]:
    """Henter excel-mapping"""
    if not excel_mappings:
        raise ValueError("excel-mapping er ikke indlæst, brug load_excel_mapping først")
    return excel_mappings


def load_excel_mapping(file_path: str, mapping_type: str = "excel"):
    """
    Indlæser excel-mapping fra en fil og gemmer den i den tilsvarende globale mapping.
    Hver workbook (worksheet) gemmes som en liste af rækker (dictionaries).

    Args:
        file_path: Stien til Excel-filen
        mapping_type: Type af mapping ("excel")
    """
    global excel_mappings

    try:
        # Load workbook
        workbook = load_workbook(file_path)

        # Initialize mapping dictionary for all worksheets
        all_sheets_mapping: Dict[str, List[Dict[str, str]]] = {}

        # Process each worksheet
        for sheet_name in workbook.sheetnames:
            worksheet = workbook[sheet_name]

            # Get header row (row 1)
            header_row = worksheet[1]
            headers = []
            for cell in header_row:
                if cell.value and str(cell.value).strip():
                    headers.append(str(cell.value).strip())

            # Initialize list for rows
            rows = []

            # Process each data row (starting from row 2)
            for row in worksheet.iter_rows(min_row=2, values_only=True):
                # Create dictionary for this row
                row_dict = {}
                for idx, header in enumerate(headers):
                    if idx < len(row):
                        cell_value = row[idx]
                        if cell_value is not None:
                            row_dict[header] = str(cell_value).strip()
                        else:
                            row_dict[header] = ""
                    else:
                        row_dict[header] = ""

                # Only add row if it has at least one non-empty value
                if any(value for value in row_dict.values()):
                    rows.append(row_dict)

            # Add this sheet's rows to the overall mapping
            all_sheets_mapping[sheet_name] = rows

        # Assign to the appropriate global mapping
        if mapping_type == "excel":
            excel_mappings = all_sheets_mapping
        else:
            raise ValueError(f"Unknown mapping_type: {mapping_type}. Expected 'excel'")

    except Exception as e:
        raise RuntimeError(
            f"Failed to load mapping from Excel file '{file_path}': {str(e)}"
        ) from e
